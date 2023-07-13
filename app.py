import os.path
from threading import Thread
import threading
from time import sleep
import PySimpleGUI as sg
import pandas as pd
import requests
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ce
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager


# configuração do driver e do wait
def start_driver():
    chrome_options = Options()
    arguments = ['--lang=pt-BR', '--start-minimized', '--headless']
    for argument in arguments:
        chrome_options.add_argument(argument)

    chrome_options.add_experimental_option('prefs', {
        # Desabilitar a confirmação de download
        'download.prompt_for_download': False,
        # Desabilitar notificações
        'profile.default_content_setting_values.notifications': 2,
        # Permitir multiplos downloads
        'profile.default_content_setting_values.automatic_downloads': 1,
    })
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])

    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=chrome_options)
    # Precisamos definir os parâmetros dentro da função
    wait = WebDriverWait(
        driver,
        10,
        poll_frequency=1,  # em quanto tempo ele vai tentar
        ignored_exceptions=[
            NoSuchElementException,
            ElementNotVisibleException,
            ElementNotSelectableException
        ]
    )

    return driver, wait


CHAVE = 'rb*8bb8$x^)w+(xxbg87s2(+6x@+x&dc)d&n((f6g@ljd&7)4g'


def get_site(site):
    site = site.replace(' ', '+')
    url = f"https://www.google.com/search?q={site}"
    driver, wait = start_driver()
    try:
        driver.get(url)
        sleep(2)
        try:
            botao_recusar = driver.find_element(By.ID, 'W0wltc')
            botao_recusar.click()
        except Exception as e:
            pass
        sleep(2)
        # link_site = driver.find_element(By.XPATH, '//*[@id="rso"]/div[1]/div/div/div/div/div/div/div[1]/a')
        # link_site = link_site.get_attribute('href')
        div_link = driver.find_element(By.CLASS_NAME, 'yuRUbf')
        link_site = div_link.find_element(By.TAG_NAME, 'a')
        link_site = link_site.get_attribute('href')
        # print(url)
        # div_link = driver.find_element(By.XPATH, '//div[@class="yuRUbf"]')
        driver.close()
        return link_site
    except Exception as e:
        link_site = 'Não encontrado'
        return link_site
    finally:
        driver.quit()


def get_instagram(instagram_in):
    instagram_in = instagram_in.replace(' ', '+')
    url = f"https://www.google.com/search?q={instagram_in}++site%3A+Instagram&ei="
    driver, wait = start_driver()
    try:
        driver.get(url)
        sleep(2)
        try:
            botao_recusar = driver.find_element(By.ID, 'W0wltc')
            botao_recusar.click()
        except Exception as e:
            pass
        sleep(2)
        div_link = driver.find_element(By.CLASS_NAME, 'yuRUbf')
        link_site = div_link.find_element(By.TAG_NAME, 'a')
        link_instagram = link_site.get_attribute('href')
        driver.close()
        return link_instagram

    except Exception as e:
        link_instagram = 'Não encontrado'
        return link_instagram
    finally:
        driver.quit()


def get_twitter(twitter_in):
    twitter_in = twitter_in.replace(' ', '+')
    url = f"https://www.google.com/search?q={twitter_in}++site%3A+twitter&ei="
    driver, wait = start_driver()
    try:
        driver.get(url)
        sleep(2)
        try:
            botao_recusar = driver.find_element(By.ID, 'W0wltc')
            botao_recusar.click()
        except Exception as e:
            pass
        sleep(2)
        div_link = driver.find_element(By.CLASS_NAME, 'yuRUbf')
        link_site = div_link.find_element(By.TAG_NAME, 'a')
        link_twitter = link_site.get_attribute('href')
        return link_twitter
    except Exception as e:
        link_twitter = 'Não encontrado'
        return link_twitter
    finally:
        driver.quit()


def get_facebook(facebook_in):
    facebook_in = facebook_in.replace(' ', '+')
    url = f"https://www.google.com/search?q={facebook_in}++site%3A+Facebook&ei="
    driver, wait = start_driver()
    try:
        driver.get(url)
        sleep(2)
        try:
            botao_recusar = driver.find_element(By.ID, 'W0wltc')
            botao_recusar.click()
        except Exception as e:
            pass
        sleep(2)
        div_link = driver.find_element(By.CLASS_NAME, 'yuRUbf')
        link_site = div_link.find_element(By.TAG_NAME, 'a')
        link_facebook = link_site.get_attribute('href')
        return link_facebook
    except Exception as e:
        link_facebook = 'Não encontrado'
        return link_facebook
    finally:
        driver.quit()


def get_youtube(youtube_in):
    youtube_in = youtube_in.replace(' ', '+')
    url = f"https://www.google.com/search?q={youtube_in}+site+%3A+YouTube&ei="
    driver, wait = start_driver()
    try:
        driver.get(url)
        sleep(2)
        try:
            botao_recusar = driver.find_element(By.ID, 'W0wltc')
            botao_recusar.click()
        except Exception as e:
            pass
        sleep(2)
        div_link = driver.find_element(By.CLASS_NAME, 'yuRUbf')
        link_site = div_link.find_element(By.TAG_NAME, 'a')
        link_youtube = link_site.get_attribute('href')
        return link_youtube
    except Exception as e:
        link_youtube = 'Não encontrado'
        return link_youtube
    finally:
        driver.quit()


def get_linkedin(linkedin_in):
    linkedin_in = linkedin_in.replace(' ', '+')
    url = f"https://www.google.com/search?q={linkedin_in}+site%3A+LinkedIn&ei="
    driver, wait = start_driver()
    try:
        driver.get(url)
        sleep(2)
        try:
            botao_recusar = driver.find_element(By.ID, 'W0wltc')
            botao_recusar.click()
        except Exception as e:
            pass
        sleep(2)
        div_link = driver.find_element(By.CLASS_NAME, 'yuRUbf')
        link_site = div_link.find_element(By.TAG_NAME, 'a')
        link_linkedin = link_site.get_attribute('href')
        return link_linkedin
    except Exception as e:
        link_linkedin = 'Não encontrado'
        return link_linkedin
    finally:
        driver.quit()


# função para verificar se tem licença válida
def valida_clave():
    # endpoint de autenticação
    auth_url = 'https://phernando-license-validation.herokuapp.com/auth'

    # credenciais para autenticação
    username = 'Fernando'
    password = 'Hermes82wars!'

    # faz a requisição de autenticação
    response = requests.post(auth_url, json={'username': username, 'password': password})

    # extrai o token de acesso do corpo da resposta
    access_token = response.json()['access_token']

    # endpoint para acessar o token do id
    token_url = 'https://phernando-license-validation.herokuapp.com/token/3'

    # adiciona o token de acesso ao header da requisição
    headers = {'Authorization': f'Bearer {access_token}'}

    # faz a requisição para obter o token do id
    response = requests.get(token_url, headers=headers)

    # extrai o token do corpo da resposta
    token = response.json()['token']

    # compara o token retornado com o token guardado em uma variável
    if token == CHAVE:
        return True
    else:
        return False


# especificar o nome do arquivo Excel
filename = 'dados.xlsx'

thread_running = True


def stop_thread():
    global thread_running
    thread_running = False


def batch_search(directory, window, instagram, twitter, facebook, youtube, linkedin, redes_sociais):
    driver = None
    try:
        global thread_running
        while thread_running:
            companies = []
            # abre o arquivo de texto
            with open(directory, 'r') as arquivo:
                # percorre cada linha do arquivo
                for c in arquivo:
                    companies.append(c)

            for c in companies:
                if c == '#':
                    thread_running = False
                    window['_progress_'].update(0)
                    print('Pesquisa concluida')
                    break

                if stop_event.is_set():
                    thread_running = False
                    # encerrar a execução da função se a flag de stop estiver definida
                    break
                # fazer uma solicitação GET para a página com o CNPJ especificado
                url = f"https://cnpj.linkana.com/busca?q={c}"
                driver, wait = start_driver()
                driver.get(url)
                window['_progress_'].update(10)
                company = wait.until(
                    ce.visibility_of_element_located((By.XPATH, '//*[@id="app"]/div/main/div/div/a/div/div[1]/p[1]')))
                sleep(2)

                company.click()
                sleep(2)
                window['_progress_'].update(20)
                company_name = driver.find_element(By.XPATH, '//*[@id="app"]/div/main/div[2]/ul[1]/li[1]/p')
                company_name = company_name.text
                cadastral_company = driver.find_element(By.XPATH,
                                                        '//*[@id="app"]/div/main/div[2]/div[1]/div/h2[2]/b[2]')
                cadastral_company = cadastral_company.text
                cadastral_situation = driver.find_element(By.XPATH, '//*[@id="app"]/div/main/div[2]/ul[1]/li[3]/p')
                cadastral_situation = cadastral_situation.text
                company_size = driver.find_element(By.XPATH, '//*[@id="app"]/div/main/div[2]/ul[1]/li[4]/p')
                company_size = company_size.text
                company_state = driver.find_element(By.XPATH, '//*[@id="app"]/div/main/div[2]/ul[2]/li[6]/p')
                company_state = company_state.text
                company_city = driver.find_element(By.XPATH, '//*[@id="app"]/div/main/div[2]/ul[2]/li[5]/p')
                company_city = company_city.text
                opening_date = driver.find_element(By.XPATH, '//*[@id="app"]/div/main/div[2]/ul[1]/li[5]/p')
                opening_date = opening_date.text
                legal_nature = driver.find_element(By.XPATH, '//*[@id="app"]/div/main/div[2]/ul[1]/li[6]/p')
                legal_nature = legal_nature.text
                share_capital = driver.find_element(By.XPATH, '//*[@id="app"]/div/main/div[2]/ul[1]/li[7]/p')
                share_capital = share_capital.text
                contact = driver.find_element(By.XPATH, '//*[@id="app"]/div/main/div[2]/ul[1]/li[11]/p')
                contact = contact.text
                phone_number = driver.find_element(By.XPATH, '//*[@id="app"]/div/main/div[2]/ul[1]/li[12]/p')
                phone_number = phone_number.text
                membership = driver.find_element(By.XPATH, '//*[@id="app"]/div/main/div[2]/div[3]/ul')
                membership = membership.text
                address = driver.find_element(By.XPATH, '//*[@id="app"]/div/main/div[2]/ul[2]')
                address = address.text
                activity_number = driver.find_element(By.XPATH, '//*[@id="app"]/div/main/div[2]/div[8]/ul/li[2]')
                activity_number = activity_number.text
                window['_progress_'].update(30)
                link_site = get_site(company_name)
                window['_progress_'].update(40)

                if instagram:
                    instagram = get_instagram(c)
                    window['_progress_'].update(50)
                else:
                    instagram = ''
                if twitter:
                    twitter = get_twitter(c)
                    window['_progress_'].update(60)
                else:
                    twitter = ''
                if facebook:
                    facebook = get_facebook(c)
                    window['_progress_'].update(70)
                else:
                    facebook = ''
                if youtube:
                    youtube = get_youtube(c)
                    window['_progress_'].update(80)
                else:
                    youtube = ''
                if linkedin:
                    linkedin = get_linkedin(c)
                    window['_progress_'].update(90)
                else:
                    linkedin = ''

                if thread_running:
                    print(f'Razão Social: {company_name}')
                    print(f'CNPJ: {cadastral_company}')
                    print(f'Situação Cadastral: {cadastral_situation}')
                    print(f'Porte: {company_size}')
                    print(f'Estado: {company_state}')
                    print(f'Cidade: {company_city}')
                    print(f'Data de abertura: {opening_date}')
                    print(f'Natureza jurídica: {legal_nature}')
                    print(f'Capital social: {share_capital}')
                    print(f'Contato: {contact}')
                    print(f'Telefone: {phone_number}')
                    print(f'Quadro Societário:\n {membership}')
                    address = address.replace("\n", " ")
                    print(f'Endereço: {address}')
                    print(f'CNAE: {activity_number}')
                    print(f'Site: {link_site} \n')
                    if redes_sociais != 0:
                        print(f'Redes Sociais: \n {instagram} \n {twitter} \n {facebook} \n {youtube} \n {linkedin} \n')
                    # adicionar uma nova linha com os valores especificados
                    new_row = pd.DataFrame({'EMPRESA': [company_name],
                                            'CNPJ': [cadastral_company],
                                            'SITUAÇÃO CADASTRAL': [cadastral_situation],
                                            'TELEFONE': [phone_number],
                                            'PORTE': [company_size],
                                            'ESTADO': [company_state],
                                            'CIDADE': [company_city],
                                            'ENDEREÇO': [address],
                                            'CAPITAL SOCIAL': [share_capital],
                                            'DATA ABERTURA': [opening_date],
                                            'SOCIOS': [membership],
                                            'CONTATO': [contact],
                                            'CNAE': [activity_number],
                                            'NATUREZA JURIDICA': [legal_nature],
                                            'SITE': [link_site],
                                            'INSTAGRAM': [instagram],
                                            'TWITTER': [twitter],
                                            'FACEBOOK': [facebook],
                                            'YOUTUBE': [youtube],
                                            'LINKEDIN': [linkedin]})

                    # verificar se o arquivo já existe
                    if os.path.isfile(filename):
                        # carregar o arquivo Excel existente
                        wb = load_workbook(filename)
                        # selecionar a planilha que será usada
                        ws = wb.active
                        # obter o índice da última linha preenchida
                        last_row = ws.max_row
                        # inserir a nova linha abaixo da última linha preenchida
                        for r in dataframe_to_rows(new_row, index=False, header=False):
                            row_num = last_row + 1
                            for c_idx, cell_value in enumerate(r, 1):
                                col_letter = get_column_letter(c_idx)
                                cell = ws[f'{col_letter}{row_num}']
                                cell.value = cell_value
                        # salvar as alterações no arquivo Excel
                        wb.save(filename)
                    else:
                        # criar um novo DataFrame com as colunas especificadas
                        df = pd.DataFrame(
                            columns=['EMPRESA', 'CNPJ', 'SITUAÇÃO CADASTRAL', 'TELEFONE', 'PORTE', 'ESTADO', 'CIDADE',
                                     'ENDEREÇO',
                                     'CAPITAL SOCIAL', 'DATA ABERTURA', 'SOCIOS', 'CONTATO', 'CNAE',
                                     'NATUREZA JURIDICA', 'SITE', 'INSTAGRAM', 'TWITTER', 'FACEBOOK', 'YOUTUBE',
                                     'LINKEDIN'])
                        # adicionar a nova linha ao DataFrame
                        df = pd.concat([df, new_row], ignore_index=True)
                        # criar um novo arquivo Excel com as colunas e a nova linha
                        wb = Workbook()
                        ws = wb.active
                        # adicionar as colunas ao arquivo Excel
                        for col_num, column_title in enumerate(df.columns, 1):
                            col_letter = get_column_letter(col_num)
                            cell = ws[f'{col_letter}1']
                            cell.value = column_title
                            cell.font = Font(bold=True)
                        # adicionar a nova linha abaixo das colunas
                        for r in dataframe_to_rows(df, index=False, header=False):
                            ws.append(r)
                        # salvar o arquivo Excel
                        wb.save(filename)
                    # input('')
                driver.close()
                window['_progress_'].update(100)
                window['_progress_'].update(0)

    except ValueError:
        print('Precisa preencher os dados corretamente')
        window['_progress_'].update(0)
    except TypeError:
        print('Precisa preencher os dados corretamente')
        window['_progress_'].update(0)
    except TimeoutException:
        print('Não encontrou dados desta empresa ou erro de Time out do servidor')
        window['_progress_'].update(0)
    finally:
        driver.quit()


def company_search(business, window, instagram, twitter, facebook, youtube, linkedin, redes_sociais):
    driver, wait = start_driver()
    try:
        # definir o CNPJ da empresa que você deseja buscar
        cnpj1 = "61.084.018/0001-03"
        cnpj2 = "62.314.844/0001-64"
        business = business.replace(' ', '%20')
        # fazer uma solicitação GET para a página com o CNPJ especificado
        url = f"https://cnpj.linkana.com/busca?q={business}"

        driver.get(url)
        window['_progress_'].update(10)
        company = wait.until(
            ce.visibility_of_element_located((By.XPATH, '//*[@id="app"]/div/main/div/div/a/div/div[1]/p[1]')))
        sleep(2)

        company.click()
        sleep(2)
        window['_progress_'].update(20)
        company_name = driver.find_element(By.XPATH, '//*[@id="app"]/div/main/div[2]/ul[1]/li[1]/p')
        company_name = company_name.text
        cadastral_company = driver.find_element(By.XPATH, '//*[@id="app"]/div/main/div[2]/div[1]/div/h2[2]/b[2]')
        cadastral_company = cadastral_company.text
        cadastral_situation = driver.find_element(By.XPATH, '//*[@id="app"]/div/main/div[2]/ul[1]/li[3]/p')
        cadastral_situation = cadastral_situation.text
        company_size = driver.find_element(By.XPATH, '//*[@id="app"]/div/main/div[2]/ul[1]/li[4]/p')
        company_size = company_size.text
        company_state = driver.find_element(By.XPATH, '//*[@id="app"]/div/main/div[2]/ul[2]/li[6]/p')
        company_state = company_state.text
        company_city = driver.find_element(By.XPATH, '//*[@id="app"]/div/main/div[2]/ul[2]/li[5]/p')
        company_city = company_city.text
        opening_date = driver.find_element(By.XPATH, '//*[@id="app"]/div/main/div[2]/ul[1]/li[5]/p')
        opening_date = opening_date.text
        legal_nature = driver.find_element(By.XPATH, '//*[@id="app"]/div/main/div[2]/ul[1]/li[6]/p')
        legal_nature = legal_nature.text
        share_capital = driver.find_element(By.XPATH, '//*[@id="app"]/div/main/div[2]/ul[1]/li[7]/p')
        share_capital = share_capital.text
        contact = driver.find_element(By.XPATH, '//*[@id="app"]/div/main/div[2]/ul[1]/li[11]/p')
        contact = contact.text
        phone_number = driver.find_element(By.XPATH, '//*[@id="app"]/div/main/div[2]/ul[1]/li[12]/p')
        phone_number = phone_number.text
        membership = driver.find_element(By.XPATH, '//*[@id="app"]/div/main/div[2]/div[3]/ul')
        membership = membership.text
        address = driver.find_element(By.XPATH, '//*[@id="app"]/div/main/div[2]/ul[2]')
        address = address.text
        activity_number = driver.find_element(By.XPATH, '//*[@id="app"]/div/main/div[2]/div[8]/ul/li[2]')
        activity_number = activity_number.text
        window['_progress_'].update(30)
        link_site = get_site(company_name)
        window['_progress_'].update(40)

        if instagram:
            instagram = get_instagram(business)
            window['_progress_'].update(50)
        else:
            instagram = ''
        if twitter:
            twitter = get_twitter(business)
            window['_progress_'].update(60)
        else:
            twitter = ''
        if facebook:
            facebook = get_facebook(business)
            window['_progress_'].update(70)
        else:
            facebook = ''
        if youtube:
            youtube = get_youtube(business)
            window['_progress_'].update(80)
        else:
            youtube = ''
        if linkedin:
            linkedin = get_linkedin(business)
            window['_progress_'].update(90)
        else:
            linkedin = ''

        print(f'Razão Social: {company_name}')
        print(f'CNPJ: {cadastral_company}')
        print(f'Situação Cadastral: {cadastral_situation}')
        print(f'Porte: {company_size}')
        print(f'Estado: {company_state}')
        print(f'Cidade: {company_city}')
        print(f'Data de abertura: {opening_date}')
        print(f'Natureza jurídica: {legal_nature}')
        print(f'Capital social: {share_capital}')
        print(f'Contato: {contact}')
        print(f'Telefone: {phone_number}')
        print(f'Quadro Societário:\n {membership}')
        address = address.replace("\n", " ")
        print(f'Endereço: {address}')
        print(f'CNAE: {activity_number}')
        print(f'Site: {link_site} \n')
        if redes_sociais != 0:
            print(f'Redes Sociais: \n {instagram} \n {twitter} \n {facebook} \n {youtube} \n {linkedin} \n')
        # adicionar uma nova linha com os valores especificados
        new_row = pd.DataFrame({'EMPRESA': [company_name],
                                'CNPJ': [cadastral_company],
                                'SITUAÇÃO CADASTRAL': [cadastral_situation],
                                'TELEFONE': [phone_number],
                                'PORTE': [company_size],
                                'ESTADO': [company_state],
                                'CIDADE': [company_city],
                                'ENDEREÇO': [address],
                                'CAPITAL SOCIAL': [share_capital],
                                'DATA ABERTURA': [opening_date],
                                'SOCIOS': [membership],
                                'CONTATO': [contact],
                                'CNAE': [activity_number],
                                'NATUREZA JURIDICA': [legal_nature],
                                'SITE': [link_site],
                                'INSTAGRAM': [instagram],
                                'TWITTER': [twitter],
                                'FACEBOOK': [facebook],
                                'YOUTUBE': [youtube],
                                'LINKEDIN': [linkedin]})

        # verificar se o arquivo já existe
        if os.path.isfile(filename):
            # carregar o arquivo Excel existente
            wb = load_workbook(filename)
            # selecionar a planilha que será usada
            ws = wb.active
            # obter o índice da última linha preenchida
            last_row = ws.max_row
            # inserir a nova linha abaixo da última linha preenchida
            for r in dataframe_to_rows(new_row, index=False, header=False):
                row_num = last_row + 1
                for c_idx, cell_value in enumerate(r, 1):
                    col_letter = get_column_letter(c_idx)
                    cell = ws[f'{col_letter}{row_num}']
                    cell.value = cell_value
            # salvar as alterações no arquivo Excel
            wb.save(filename)
        else:
            # criar um novo DataFrame com as colunas especificadas
            df = pd.DataFrame(
                columns=['EMPRESA', 'CNPJ', 'SITUAÇÃO CADASTRAL', 'TELEFONE', 'PORTE', 'ESTADO', 'CIDADE', 'ENDEREÇO',
                         'CAPITAL SOCIAL', 'DATA ABERTURA', 'SOCIOS', 'CONTATO', 'CNAE', 'NATUREZA JURIDICA', 'SITE',
                         'INSTAGRAM', 'TWITTER', 'FACEBOOK', 'YOUTUBE', 'LINKEDIN'])
            # adicionar a nova linha ao DataFrame
            df = pd.concat([df, new_row], ignore_index=True)
            # criar um novo arquivo Excel com as colunas e a nova linha
            wb = Workbook()
            ws = wb.active
            # adicionar as colunas ao arquivo Excel
            for col_num, column_title in enumerate(df.columns, 1):
                col_letter = get_column_letter(col_num)
                cell = ws[f'{col_letter}1']
                cell.value = column_title
                cell.font = Font(bold=True)
            # adicionar a nova linha abaixo das colunas
            for r in dataframe_to_rows(df, index=False, header=False):
                ws.append(r)
            # salvar o arquivo Excel
            wb.save(filename)
            # input('')
            driver.close()

    except ValueError:
        print('Precisa preencher os dados corretamente')
    except TypeError:
        print('Precisa preencher os dados corretamente')
    except TimeoutException:
        print('Não encontrou dados desta empresa ou erro de Time out do servidor')
    finally:
        driver.quit()
    print('Pesquisa concluida')
    window['_progress_'].update(100)
    window['_progress_'].update(0)


sg.theme('SystemDefaultForReal')
# layout das colunas

coluna_output = [
    [sg.Output(size=(110, 24), key='_output_')],
    [sg.ProgressBar(max_value=100, orientation='h', size=(60, 20), key='_progress_')]
]

coluna_esquerda = [
    [sg.Text(size=(100, 1)), sg.Text(size=(10, 1))],
    [sg.Text('Razão social ou CNPJ')],
    [sg.Input(size=(100, 0), key='company')],
    [sg.Checkbox("Buscar em lote", default=True, key='lote'), sg.Checkbox("Instagram", default=True, key='instagram'),
     sg.Checkbox("Twitter", default=True, key='twitter'), sg.Checkbox("Facebook", default=True, key='facebook'),
     sg.Checkbox("YouTube", default=True, key='youtube'), sg.Checkbox("LinkedIn", default=True, key='linkedin')],
    [sg.Text(size=(100, 1)), sg.Text(size=(10, 1))],
    [sg.Text("Lista de empresas: "), sg.Input(size=(73, 1), key='directory'), sg.FileBrowse()],
]

# layout principal
layout_principal = [
    [sg.Frame('Parâmetros de busca', coluna_esquerda)],
    [sg.Column(coluna_output)],
    [sg.Button('Start', key='execute'), sg.Button('Stop  ', key='stop')],
]

if __name__ == '__main__':
    # janela
    window = sg.Window('Busca Empresas', layout_principal, finalize=True, icon="icone.ico", return_keyboard_events=True,
                       size=(800, 680))

    stop_event = threading.Event()

    while True:
        try:
            event, values = window.read()

            if event == sg.WIN_CLOSED:
                break

            if event == 'stop':
                window.FindElement('_output_').Update('')
                stop_event.set()
                thread_stop = Thread(target=stop_thread, args=(), daemon=True)
                thread_stop.start()
                print('Aplicação encerrada!')
                sleep(15)
                window['_progress_'].update(0)

            if event == 'execute' or event == '\r':  # '\r' representa a tecla Enter
                window.FindElement('_output_').Update('')
                # if valida_clave():
                lote = False
                business = values['company']
                instagram = True
                twitter = True
                facebook = True
                youtube = True
                linkedin = True
                redes_sociais: int = 5
                if not values['instagram']:
                    instagram = False
                    redes_sociais -= 1
                if not values['twitter']:
                    twitter = False
                    redes_sociais -= 1
                if not values['facebook']:
                    facebook = False
                    redes_sociais -= 1
                if not values['youtube']:
                    youtube = False
                    redes_sociais -= 1
                if not values['linkedin']:
                    linkedin = False
                    redes_sociais -= 1
                if values['lote'] and not values['directory']:
                    print('Precisa selecionar o arquivo com a lista de empresas')

                elif values['lote'] and business:
                    print('Inciando presquisa em lote... ')
                    business = ''
                    lote = True
                    window['_progress_'].update(0)
                    directory = values['directory']
                    thread_catch_batch = Thread(target=batch_search, args=(
                        directory, window, instagram, twitter, facebook, youtube, linkedin, redes_sociais),
                                                daemon=True)
                    thread_catch_batch.start()
                    print('Buscando dados...\n')

                elif values['lote'] and not business:
                    lote = True
                    business = ''
                    window['_progress_'].update(0)
                    directory = values['directory']
                    thread_catch_batch = Thread(target=batch_search, args=(
                        directory, window, instagram, twitter, facebook, youtube, linkedin, redes_sociais),
                                                daemon=True)
                    thread_catch_batch.start()
                    print('Buscando dados...\n')

                else:
                    window['_progress_'].update(0)
                    thread_catch_company = Thread(target=company_search, args=(
                        business, window, instagram, twitter, facebook, youtube, linkedin, redes_sociais),
                                                  daemon=True)
                    thread_catch_company.start()
                    print('Buscando dados...\n')
                # else:
                #     print('Erro no servidor Heroku')

        except Exception as e:
            print(f'Erro na execução: \n {e}')
            break

    # Fecha a janela e encerra a aplicação
    window.close()
